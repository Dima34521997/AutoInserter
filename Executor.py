import openpyxl
from datetime import date

from openpyxl.styles import (Alignment,
                             Border,
                             Side)


def get_msl_numbers_list(ds):
    pass


def execute(ds):
    wb = openpyxl.load_workbook(ds.table)
    ws = wb.active

    # ws = wb.get_sheet_by_name('ПИ 026-04')

    free_cell = 2

    while ws[f'A{free_cell}'].value is not None:
        free_cell += 1

    try:
        ds.msl_number = int(ws[f'A{free_cell - 1}'].value) + 1
    except ValueError:
        pass

    chunk = ds.amount // ds.per_one_msl
    residue = ds.amount % ds.per_one_msl

    borders = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    if residue == 0:
        for i in range(chunk):
            ws[f'A{free_cell + i}'] = ds.msl_number + i
            ws[f'A{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'A{free_cell + i}'].border = borders

            ws[f'B{free_cell + i}'] = ds.device_name
            ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'B{free_cell + i}'].border = borders

            ws[f'C{free_cell + i}'] = ds.per_one_msl
            ws[f'C{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'C{free_cell + i}'].border = borders

            if ds.first - (ds.first + ds.per_one_msl - 1) != 0:
                ws[f'D{free_cell + i}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
            else:
                ws[f'D{free_cell + i}'] = f'{ds.first}'
            # ws[f'D{free_cell + i}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
            # ws[f'D{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'D{free_cell + i}'].border = borders

            ws[f'E{free_cell + i}'] = date.today()
            ws[f'E{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'E{free_cell + i}'].border = borders

            ws[f'F{free_cell + i}'] = ds.master_name
            ws[f'F{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'F{free_cell + i}'].border = borders

            ws[f'G{free_cell + i}'] = ds.contract
            ws[f'G{free_cell + i}'].alignment = Alignment(horizontal='center')
            ws[f'G{free_cell + i}'].border = borders

            # region Maxmod
            # ws[f'A{free_cell + i}'] = ds.msl_number + i
            # ws[f'A{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'A{free_cell + i}'].border = borders
            #
            # # ws[f'B{free_cell + i}'] = ds.device_name
            # # ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')
            # # ws[f'B{free_cell + i}'].border = borders
            #
            # ws[f'B{free_cell + i}'] = ds.per_one_msl
            # ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'B{free_cell + i}'].border = borders
            #
            # ws[f'C{free_cell + i}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
            # ws[f'C{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'C{free_cell + i}'].border = borders
            #
            # ws[f'D{free_cell + i}'] = date.today()
            # ws[f'D{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'D{free_cell + i}'].border = borders
            #
            # ws[f'E{free_cell + i}'] = ds.master_name
            # ws[f'E{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'E{free_cell + i}'].border = borders
            #
            # ws[f'F{free_cell + i}'] = ds.contract
            # ws[f'F{free_cell + i}'].alignment = Alignment(horizontal='center')
            # ws[f'F{free_cell + i}'].border = borders
            # endregion

            ds.first += ds.per_one_msl

    else:
        count = 0
        for count in range(chunk + 1):
            ws[f'A{free_cell + count}'] = ds.msl_number + count
            ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'A{free_cell + count}'].border = borders

            ws[f'B{free_cell + count}'] = ds.device_name
            ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'B{free_cell + count}'].border = borders

            ws[f'C{free_cell + count}'] = ds.per_one_msl
            ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'C{free_cell + count}'].border = borders

            ws[f'D{free_cell + count}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
            ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'D{free_cell + count}'].border = borders

            ws[f'E{free_cell + count}'] = date.today().strftime("%d.%m.%Y")
            ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'E{free_cell + count}'].border = borders

            ws[f'F{free_cell + count}'] = ds.master_name
            ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'F{free_cell + count}'].border = borders

            ws[f'G{free_cell + count}'] = ds.contract
            ws[f'G{free_cell + count}'].alignment = Alignment(horizontal='center')
            ws[f'G{free_cell + count}'].border = borders

            # region Maxmod
            # ws[f'A{free_cell + count}'] = ds.msl_number + count
            # ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'A{free_cell + count}'].border = borders
            #
            # # ws[f'B{free_cell + i}'] = ds.device_name
            # # ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')
            # # ws[f'B{free_cell + i}'].border = borders
            #
            # ws[f'B{free_cell + count}'] = ds.per_one_msl
            # ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'B{free_cell + count}'].border = borders
            #
            # ws[f'C{free_cell + count}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
            # ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'C{free_cell + count}'].border = borders
            #
            # ws[f'D{free_cell + count}'] = date.today()
            # ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'D{free_cell + count}'].border = borders
            #
            # ws[f'E{free_cell + count}'] = ds.master_name
            # ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'E{free_cell + count}'].border = borders
            #
            # ws[f'F{free_cell + count}'] = ds.contract
            # ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')
            # ws[f'F{free_cell + count}'].border = borders
            # endregion

            ds.first += ds.per_one_msl

        ws[f'A{free_cell + count}'] = ds.msl_number + count
        ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'A{free_cell + count}'].border = borders

        ws[f'B{free_cell + count}'] = ds.device_name
        ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'B{free_cell + count}'].border = borders

        ws[f'C{free_cell + count}'] = residue
        ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'C{free_cell + count}'].border = borders
        kurwa = ds.first - ds.per_one_msl - ds.first + residue + ds.per_one_msl - 1
        if kurwa != 0:
            ws[f'D{free_cell + count}'] = f'{ds.first - ds.per_one_msl} - {ds.first + residue - ds.per_one_msl - 1}'
        else:
            ws[f'D{free_cell + count}'] = f'{ds.first - ds.per_one_msl}'

        ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'D{free_cell + count}'].border = borders

        ws[f'E{free_cell + count}'] = date.today().strftime("%d.%m.%Y")
        ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'E{free_cell + count}'].border = borders

        ws[f'F{free_cell + count}'] = ds.master_name
        ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'F{free_cell + count}'].border = borders

        ws[f'G{free_cell + count}'] = ds.contract
        ws[f'G{free_cell + count}'].alignment = Alignment(horizontal='center')
        ws[f'G{free_cell + count}'].border = borders

        # region Maxmod
        # ws[f'A{free_cell + count}'] = ds.msl_number + count
        # ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'A{free_cell + count}'].border = borders
        #
        # # ws[f'B{free_cell + i}'] = ds.device_name
        # # ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')
        # # ws[f'B{free_cell + i}'].border = borders
        #
        # ws[f'B{free_cell + count}'] = ds.per_one_msl
        # ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'B{free_cell + count}'].border = borders
        #
        # ws[f'C{free_cell + count}'] = f'{ds.first} - {ds.first + ds.per_one_msl - 1}'
        # ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'C{free_cell + count}'].border = borders
        #
        # ws[f'D{free_cell + count}'] = date.today()
        # ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'D{free_cell + count}'].border = borders
        #
        # ws[f'E{free_cell + count}'] = ds.master_name
        # ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'E{free_cell + count}'].border = borders
        #
        # ws[f'F{free_cell + count}'] = ds.contract
        # ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')
        # ws[f'F{free_cell + count}'].border = borders
        # endregion

    wb.save(ds.table)
    return True


